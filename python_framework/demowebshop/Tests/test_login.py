#import Pytest as Pytest
import allure
import pytest
import openpyxl
from allure_commons.types import AttachmentType

from utilities.excel_reader import get_list_from_excel

from demowebshop.POM.homepage import HomePage
from demowebshop.POM.loginpage import LogIn
#usernames=[("tabjulmanju@gmail.com","Manju@468"),("abcdd@gmail.com","123"),("xyzz@gmail.com","233"),("zzz@gmail.com","111")]
#@pytest.mark.parametrize("password",["123","999","888"])

# creds=openpyxl.load_workbook("threadmanju.xlsx")
# cred_login=creds["cred_login"]
# #login_creds.cell(1,2)
# credentials = []
# for row in range(1,3):
#     nested_creds = []
#     for col in range(1,3):
#         data = cred_login.cell(row,col)
#         nested_creds.append(data.value)
#     credentials.append(nested_creds)
#print(credentials)

credentials=get_list_from_excel("threadmanju.xlsx","cred_login")

@pytest.mark.parametrize("username,password",credentials)
def test_login_with_proper_credentials(driver,username,password):
    home = HomePage(driver)
    home.click_on_login()
    login=LogIn(driver)
    login.login_to_the_application(username,password)
    condition=driver.find_element("xpath",f"//a[.='{username}']").is_displayed()
    #condition = driver.find_element("xpath", f"//a[.='{username}']").is_displayed()
    if not condition:
        allure.attach(driver.get_screenshot_as_png(),name="test_login_with_proper_credentials.png",attachment_type=AttachmentType.PNG)
    assert condition
    #assert false--->false


# @pytest.mark.skip
# def test_login_with_no_password(driver):
#     home = HomePage(driver)
#     home.click_on_login()
#     login = LogIn(driver)
#     login.login_to_the_application("tabjulmanju@gmail.com","")
#     #assert driver.find_element("xpath", "//a[.='tabjulmanju@gmail.com']").is_displayed()
#     driver.find_element("xpath","//span[contains(.,'Login was unsuccessful.')]").is_displayed()
# @pytest.mark.skip
# def test_login_with_no_username(driver):
#     home = HomePage(driver)
#     home.click_on_login()
#     login = LogIn(driver)
#     login.login_to_the_application("","Manju@468")
#     #assert driver.find_element("xpath", "//a[.='tabjulmanju@gmail.com']").is_displayed()
#     driver.find_element("xpath","//span[contains(.,'Login was unsuccessful.')]").is_displayed()
# @pytest.mark.skip
# def test_login_with_no_credentials(driver):
#     home = HomePage(driver)
#     home.click_on_login()
#     login = LogIn(driver)
#     login.login_to_the_application("","")
#     #assert driver.find_element("xpath", "//a[.='tabjulmanju@gmail.com']").is_displayed()
#     driver.find_element("xpath","//span[contains(.,'Login was unsuccessful.')]").is_displayed()
# @pytest.mark.skip
# def test_login_with_invalid_Credentials(driver):
#     home = HomePage(driver)
#     home.click_on_login()
#     login = LogIn(driver)
#     login.login_to_the_application("tabjulmanju@gmail.com", "Manjus@468")
#     #assert driver.find_element("xpath", "//a[.='tabjulmanju@gmail.com']").is_displayed()
#     driver.find_element("xpath", "//span[contains(.,'Login was unsuccessful.')]").is_displayed()
#
#

@pytest.mark.xfail
def test_sample1():
    print("Expected Pass")
@pytest.mark.xfail
def test_sample2():
    print("Executing sample2")
    assert True
@pytest.mark.xfail
def test_sample3():
    print("Expected fail")
    assert False