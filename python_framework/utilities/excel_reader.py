import openpyxl

# def get_list_from_excel():
#     creds=openpyxl.load_workbook("threadmanju.xlsx")
#     cred_login=creds["cred_login"]
#     #login_creds.cell(1,2)
#     credentials = []
#     for row in range(1,3):
#         nested_creds = []
#         for col in range(1,3):
#             data = cred_login.cell(row,col).value
#             nested_creds.append(data)
#         credentials.append(nested_creds)
#     return credentials
def get_list_from_excel(excel_name,sheet_name):
    excel=openpyxl.load_workbook(excel_name)
    sheet=excel[sheet_name]
    #login_creds.cell(1,2)
    values = []
    for row in range(1,sheet.max_row+1):
        nested_creds = []
        for col in range(1,sheet.max_column+1):
            data = sheet.cell(row,col)
            nested_creds.append(data.value)
        values.append(nested_creds)
    return values

def add_value_to_excel(excel_name,sheet_name):
    excel = openpyxl.load_workbook(excel_name)
    sheet = excel[sheet_name]
    sheet.cell(2,1).value="new value"
    excel.save(excel_name)

add_value_to_excel("./threadmanju.xlsx","cred_login")